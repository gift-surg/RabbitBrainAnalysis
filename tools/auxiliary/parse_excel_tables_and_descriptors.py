import os

import numpy as np
import openpyxl as oxl


def parse_excel_data_to_list(pfi_excel_file, worksheet_name, min_col=0, max_col=12):

    if not os.path.exists(pfi_excel_file):
        raise IOError('Input file {} does not exists'.format(pfi_excel_file))
    if not (pfi_excel_file.endswith('.xlsx') or pfi_excel_file.endswith('.xls')):
        raise IOError('Input file {} is not an excel file')

    wb = oxl.load_workbook(pfi_excel_file)
    ws = wb[worksheet_name]

    ans = []
    for row in ws.iter_rows(min_col=min_col, max_col=max_col,):
        ans_row = []
        for cell in row:
            ans_row.append(cell.value)
        ans.append(ans_row)

    for l_row in ans[1:]:
        l_row[0] = str(l_row[0]).replace('.', '')

    return ans


def parse_label_descriptor_in_a_list(pfi_label_descriptor):
    """
    parse the ITK-Snap into a list.
    :param pfi_label_descriptor: path to file to label descriptor.
    :return: list of lists, where each sublist contains the information of each line.
    """
    if not os.path.exists(pfi_label_descriptor):
        msg = 'Label descriptor file {} does not exist'.format(pfi_label_descriptor)
        raise IOError(msg)

    f = open(pfi_label_descriptor, 'r')
    lines = f.readlines()

    label_descriptor_list = []

    for l in lines:
        if not l.startswith('#'):

            parsed_line = [j.strip() for j in l.split('  ') if not j == '']
            for position_element, element in enumerate(parsed_line):
                if element.isdigit():
                    parsed_line[position_element] = int(element)
                if element.startswith('"') or element.endswith('"'):
                    parsed_line[position_element] = element.replace('"', '')

            parsed_line.insert(1, parsed_line[-1])

            label_descriptor_list.append(parsed_line[:-1])

    return label_descriptor_list


def parse_multi_label_descriptor_in_a_list(pfi_multi_lab_descriptor):
    """
    Parse the multi-label descriptor (created on the idea of the label descriptor) in a list.
    :param pfi_multi_lab_descriptor:
    :return:
    """
    if not os.path.exists(pfi_multi_lab_descriptor):
        msg = 'Label descriptor file {} does not exist'.format(pfi_multi_lab_descriptor)
        raise IOError(msg)

    f = open(pfi_multi_lab_descriptor, 'r')
    lines = f.readlines()

    label_descriptor_list = []

    for l in lines:
        if not l.startswith('#'):
            parsed_line = [j.strip() for j in l.split('&')]
            for k in xrange(1, len(parsed_line)):
                parsed_line[k] = int(parsed_line[k])
            label_descriptor_list.append(parsed_line)

    return label_descriptor_list


def write_header_excel_tabs_from_record(pfi_excel, sheet_name, pfi_record):

    assert os.path.exists(pfi_record)
    assert os.path.exists(pfi_excel)

    wb = oxl.load_workbook(pfi_excel)

    assert sheet_name in wb.get_sheet_names()

    ws = wb.get_sheet_by_name(sheet_name)

    record_dict = np.load(pfi_record)
    record_dict = record_dict.item()
    regions = record_dict['Regions'][1:-1]

    del record_dict  # save space

    fixed_header = ['Age', 'ID Number', 'Delivery Gestation (g)', 'Weight PND1 (g)', 'Harvest Date',
                    'Brain Weight (g)', 'Brain Volume (ml)', 'Sex', 'MRI Date', 'Acquisition', 'Invivo MRI isofluorane time',
                    'MRI Number', 'MRI Acquisition Format', 'Brain Volume T1', 'Brain Volume FA', 'Brain Volume ADC',
                    'Brain Volume g-ratio', 'ICV', 'A-P', 'S-I', 'R-L']
    T1_header = [reg + ' Vol ' for reg in regions]
    FA_header = [reg + ' FA ' for reg in regions]
    ADC_header = [reg + ' ADC ' for reg in regions]
    g_ratio_header = [reg + ' g-ratio ' for reg in regions]

    hd = fixed_header + T1_header + FA_header + ADC_header + g_ratio_header

    for col, col_val in enumerate(hd):
        ws.cell(row=1, column=col+1).value = hd[col]

    wb.save(pfi_excel)


def store_a_record_in_excel_table(pfi_record, pfi_excel, sj, sheet_name):

    # Very deliccate and very not robust... Assert will be everywhere to increase robustness!
    assert os.path.exists(pfi_record)
    assert os.path.exists(pfi_excel)
    # load workbook and work sheet
    wb = oxl.load_workbook(pfi_excel)
    assert sheet_name in wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheet_name)
    # get excel file header
    first_row = ws.rows.next()
    first_row_starting_header = [g.value for g in first_row[:33]]
    # get the column with all the ids and store it in a list in a list
    id_numbers_col_num = first_row_starting_header.index('ID Number')
    row_ids = ws.iter_rows(min_row=2, min_col=id_numbers_col_num + 1, max_col=id_numbers_col_num + 1)
    id_nums = []
    for p in row_ids:
        if p[0].value is not None:
            id_nums.append(str(p[0].value).replace('.', ''))
    # get the first column num of the row we want to populate
    starting_col_num = first_row_starting_header.index('Brain Volume T1')
    assert first_row_starting_header[starting_col_num] == 'Brain Volume T1'
    assert sj in id_nums
    # get row where the given id is stored
    row_num_id = id_nums.index(sj)
    assert str(ws.cell(row=row_num_id+2, column=id_numbers_col_num+1).value).replace('.', '') == sj
    del first_row, first_row_starting_header, id_numbers_col_num, row_ids, id_nums, p  # spare space
    # At this point we have the row where we want to write the stuff and the col from where we want to start writing it
    # print row_num_id, starting_col_num
    # now we need to store the stuff to write in a list:
    record_dict = np.load(pfi_record)
    record_dict = record_dict.item()
    vols_T1 = list(record_dict['vols'][1:-1])
    vals_FA = list(record_dict['FAs'][1:-1])
    vals_ADC = list(record_dict['ADCs'][1:-1])
    vals_g_ratio = list(record_dict['g_ratios'][1:-1])
    col_vals = [record_dict['Info']['totVol T1'], record_dict['Info']['totVol FA'], record_dict['Info']['totVol ADC'],
                record_dict['Info']['totVol g_ratio'], record_dict['Info']['ICV']]
    col_vals += vols_T1 + vals_FA + vals_ADC + vals_g_ratio
    # ...and to write it in the appropriate place:
    for col_num, col_val in enumerate(col_vals):
        ws.cell(column=starting_col_num + col_num + 1, row=row_num_id + 2).value = col_val

    #
    # k = 0
    # for p in cells_to_fill:
    #     p[0].value = row_vals[k]
    #     k += 1
    # finally save the excel file
    wb.save(pfi_excel)


if __name__ == '__main__':

    # update the header with the latests records.
    from tools.definitions import root_study_rabbits

    pfi_record_for_a_subject = os.path.join(root_study_rabbits, 'A_data', 'PTB', 'ex_vivo', '1201', 'records',
                                            '1201_record.npy')
    pfi_excel_file = os.path.join(root_study_rabbits, 'A_data', 'Utils', 'REoP_Data.xlsx')
    if 1:
        write_header_excel_tabs_from_record(pfi_excel_file, 'ACS', pfi_record_for_a_subject)
        write_header_excel_tabs_from_record(pfi_excel_file, 'PTB', pfi_record_for_a_subject)
        write_header_excel_tabs_from_record(pfi_excel_file, 'Template', pfi_record_for_a_subject)
    if 0:
        store_a_record_in_excel_table(pfi_record_for_a_subject, pfi_excel_file, '1201', 'PTB')
