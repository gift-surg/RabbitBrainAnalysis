import os
from openpyxl import load_workbook


def parse_excel_data_to_list(pfi_excel_file):

    if not os.path.exists(pfi_excel_file):
        raise IOError('Input file {} does not exists'.format(pfi_excel_file))
    if not (pfi_excel_file.endswith('.xlsx') or pfi_excel_file.endswith('.xls')):
        raise IOError('Input file {} is not an excel file')

    wb = load_workbook(pfi_excel_file)
    ws = wb.active

    ans = []
    for row in ws.iter_rows(min_col=0, max_col=8,):
        ans_row = []
        for cell in row:
            ans_row.append(cell.value)
        ans.append(ans_row)

    for l_row in ans[1:]:
        l_row[0] = str(l_row[0]).replace('.', '')

    return ans


# pfi_data = '/Users/sebastiano/Dropbox/RabbitEOP-MRI/docs/REoP_Pilot_MRI_Data.xlsx'
# parse_excel_data_to_list(pfi_data)


