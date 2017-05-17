import os
from openpyxl import load_workbook


def parse_excel_data_to_list(pfi_excel_file):

    if not os.path.exists(pfi_excel_file):
        raise IOError('Input file {} does not exists'.format(pfi_excel_file))
    if not (pfi_excel_file.endswith('.xlsx') or pfi_excel_file.endswith('.xls')):
        raise IOError('Input file {} is not an excel file')

    wb = load_workbook(pfi_excel_file)
    ws = wb.active

    ans = []
    for row in ws.iter_rows(min_col=0, max_col=8,):
        ans_row = []
        for cell in row:
            ans_row.append(cell.value)
        ans.append(ans_row)

    for l_row in ans[1:]:
        l_row[0] = str(l_row[0]).replace('.', '')

    return ans


def parse_label_descriptor_in_a_list(pfi_label_descriptor):
    """
    parse the ITK-Snap into a list.
    :param pfi_label_descriptor: path to file to label descriptor.
    :return: list of lists, where each sublist contains the information of each line.
    """
    if not os.path.exists(pfi_label_descriptor):
        msg = 'Label descriptor file {} does not exist'.format(pfi_label_descriptor)
        raise IOError(msg)

    f = open(pfi_label_descriptor, 'r')
    lines = f.readlines()

    label_descriptor_list = []

    for l in lines:
        if not l.startswith('#'):

            parsed_line = [j.strip() for j in l.split('  ') if not j == '']
            for position_element, element in enumerate(parsed_line):
                if element.isdigit():
                    parsed_line[position_element] = int(element)
                if element.startswith('"') or element.endswith('"'):
                    parsed_line[position_element] = element.replace('"', '')

            parsed_line.insert(1, parsed_line[-1])

            label_descriptor_list.append(parsed_line[:-1])

    return label_descriptor_list


def parse_multi_label_descriptor_in_a_list(pfi_multi_lab_descriptor):
    """
    Parse the multi-label descriptor (created on the idea of the label descriptor) in a list.
    :param pfi_multi_lab_descriptor:
    :return:
    """
    if not os.path.exists(pfi_multi_lab_descriptor):
        msg = 'Label descriptor file {} does not exist'.format(pfi_multi_lab_descriptor)
        raise IOError(msg)

    f = open(pfi_multi_lab_descriptor, 'r')
    lines = f.readlines()

    label_descriptor_list = []

    for l in lines:
        if not l.startswith('#'):
            parsed_line = [j.strip() for j in l.split('&')]
            label_descriptor_list.append(parsed_line)

    return label_descriptor_list


pfi_multi_lab = '/Users/sebastiano/Desktop/test_main/Utils/multi_label_descriptor.txt'
li = parse_multi_label_descriptor_in_a_list(pfi_multi_lab)
print li
